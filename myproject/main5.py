from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl

# Configuration
CHROMEDRIVER_PATH = 'chromedriver.exe'
OUTPUT_FILE = 'message_history.xlsx'
USERNAME = "your_username"
PASSWORD = "your_password"
PRECONFIGURED_MESSAGE = "Thank you very much!"

def setup_driver():
    # Set up Chrome WebDriver
    service = Service(executable_path=CHROMEDRIVER_PATH)
    driver = webdriver.Chrome(service=service)
    driver.maximize_window()
    return driver

def login_to_linkedin(driver, username, password):
    try:
        driver.get("https://www.linkedin.com/login")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'username'))).send_keys(username)
        driver.find_element(By.ID, 'password').send_keys(password)
        driver.find_element(By.ID, 'password').send_keys(Keys.RETURN)
        
        # Wait for 2FA manually and press Enter
        input("Please complete 2FA and press Enter...")
        print("LinkedIn login successful.")

    except Exception as e:
        print(f"Error during LinkedIn login: {e}")
        driver.save_screenshot('error_screenshot_login.png')

def navigate_to_messaging(driver):
    try:
        driver.get("https://www.linkedin.com/messaging/")
        WebDriverWait(driver, 20).until(EC.url_contains("messaging"))
        print("Navigated to LinkedIn messaging.")
        
    except Exception as e:
        print(f"Error navigating to messaging: {e}")
        driver.save_screenshot('error_screenshot_navigation.png')

def send_preconfigured_message(driver, sheet, history):
    try:
        # Click on the Unread button
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//button[contains(@aria-label, "Unread")]'))).click()
        time.sleep(5)  # Wait for unread messages to load

        # Find all unread messages
        unread_messages = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'li.msg-conversation-listitem')))

        for message in unread_messages:
            try:
                # Extract person's name
                person_name = message.find_element(By.CSS_SELECTOR, 'span.msg-conversation-listitem__participant-names').text
                
                # Check if message already processed
                if person_name in history:
                    continue
                
                # Click on the message to open it
                message.click()
                time.sleep(3)  # Wait for message to open

                # Find and click the preconfigured message button
                preconfigured_message_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, f'//button[contains(text(), "{PRECONFIGURED_MESSAGE}")]')))
                preconfigured_message_button.click()

                # Update message history
                update_excel(sheet, person_name, PRECONFIGURED_MESSAGE)
                time.sleep(20)  # Delay between messages

                print(f"Message sent to {person_name}.")

            except Exception as e:
                print(f"Error processing message for {person_name}: {e}")
                driver.save_screenshot(f'error_screenshot_{person_name}.png')

    except Exception as e:
        print(f"Error occurred: {e}")

def setup_excel():
    try:
        workbook = openpyxl.load_workbook(OUTPUT_FILE)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Person", "Message"])
    return workbook, sheet

def update_excel(sheet, person_name, message):
    sheet.append([person_name, message])

def save_excel(workbook):
    workbook.save(OUTPUT_FILE)

def main():
    try:
        # Set up WebDriver and login
        driver = setup_driver()
        login_to_linkedin(driver, USERNAME, PASSWORD)

        # Navigate to messaging
        navigate_to_messaging(driver)

        # Set up Excel file
        workbook, sheet = setup_excel()

        # Read history from Excel
        history = {row[0].value: row[1].value for row in sheet.iter_rows(min_row=2)}

        # Send preconfigured messages
        send_preconfigured_message(driver, sheet, history)

        # Save Excel file
        save_excel(workbook)

    except Exception as e:
        print(f"Unexpected error occurred: {e}")

    finally:
        # Close the browser
        driver.quit()

if __name__ == "__main__":
    main()
